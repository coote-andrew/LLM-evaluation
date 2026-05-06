"""
CSV/Excel ingestion with input_/output_ column convention.

Columns prefixed with input_ are fed into the prompt.
Columns prefixed with output_ are used for scoring.
"""

import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _normalize_columns(columns: list[str]) -> tuple[list[str], list[str], list[str]]:
    """Split columns into input_, output_, and other."""
    input_cols = [c for c in columns if c.startswith("input_")]
    output_cols = [c for c in columns if c.startswith("output_")]
    all_cols = list(columns)
    return all_cols, input_cols, output_cols


def parse_csv(content: bytes | str, filename: str = "") -> dict[str, Any]:
    """
    Parse CSV content. Returns dict with:
    - column_names: list of all column names
    - input_columns: columns starting with input_
    - output_columns: columns starting with output_
    - rows: list of dicts, each with input_fields and expected_output_fields
    - row_count: number of rows
    """
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(content))
    raw_columns = reader.fieldnames or []
    columns = [c.strip() for c in raw_columns]
    all_cols, input_cols, output_cols = _normalize_columns(columns)

    rows = []
    for i, raw_row in enumerate(reader, start=1):
        row = {k.strip(): v for k, v in raw_row.items() if k is not None}
        input_fields = {k: row.get(k, "") for k in input_cols if k in row}
        expected_output_fields = {k: row.get(k, "") for k in output_cols if k in row}
        rows.append({
            "row_number": i,
            "input_fields": input_fields,
            "expected_output_fields": expected_output_fields,
        })

    return {
        "column_names": all_cols,
        "input_columns": input_cols,
        "output_columns": output_cols,
        "rows": rows,
        "row_count": len(rows),
        "original_filename": filename or "upload.csv",
    }


def parse_excel(content: bytes, filename: str = "") -> dict[str, Any]:
    """
    Parse Excel (.xlsx) content. Uses first sheet.
    Returns same structure as parse_csv.
    """
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    if not sheet:
        return {
            "column_names": [],
            "input_columns": [],
            "output_columns": [],
            "rows": [],
            "row_count": 0,
            "original_filename": filename or "upload.xlsx",
        }

    rows_iter = sheet.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return {
            "column_names": [],
            "input_columns": [],
            "output_columns": [],
            "rows": [],
            "row_count": 0,
            "original_filename": filename or "upload.xlsx",
        }

    columns = [str(c).strip() if c is not None else "" for c in header]
    all_cols, input_cols, output_cols = _normalize_columns(columns)

    rows = []
    for i, row_values in enumerate(rows_iter, start=1):
        row_dict = dict(zip(columns, row_values or []))
        input_fields = {
            k: ("" if row_dict.get(k) is None else str(row_dict.get(k)))
            for k in input_cols if k in row_dict
        }
        expected_output_fields = {
            k: ("" if row_dict.get(k) is None else str(row_dict.get(k)))
            for k in output_cols if k in row_dict
        }
        rows.append({
            "row_number": i,
            "input_fields": input_fields,
            "expected_output_fields": expected_output_fields,
        })

    wb.close()
    return {
        "column_names": all_cols,
        "input_columns": input_cols,
        "output_columns": output_cols,
        "rows": rows,
        "row_count": len(rows),
        "original_filename": filename or "upload.xlsx",
    }


def group_rows(
    rows: list[dict],
    group_by_cols: list[str],
    sort_by_col: str | None = None,
) -> list[dict]:
    """
    Aggregate flat rows into one row per unique group-key.

    Static fields (group_by_cols) are hoisted to the top level of input_fields.
    All remaining input_* fields are collected into input_notes as a list of dicts.
    expected_output_fields is taken from the first row of each group.
    """
    groups: dict[tuple, list[dict]] = {}
    group_order: list[tuple] = []

    for row in rows:
        key = tuple(row["input_fields"].get(col, "") for col in group_by_cols)
        if key not in groups:
            groups[key] = []
            group_order.append(key)
        groups[key].append(row)

    result = []
    for i, key in enumerate(group_order, start=1):
        group = groups[key]

        if sort_by_col:
            group = sorted(
                group,
                key=lambda r: r["input_fields"].get(sort_by_col, "") or "",
            )

        first = group[0]
        static_fields = {col: first["input_fields"].get(col, "") for col in group_by_cols}

        note_cols = [k for k in first["input_fields"] if k not in group_by_cols]
        notes = [
            {col: row["input_fields"].get(col, "") for col in note_cols}
            for row in group
        ]

        result.append({
            "row_number": i,
            "input_fields": {**static_fields, "input_notes": notes},
            "expected_output_fields": first["expected_output_fields"],
        })

    return result


def parse_upload(
    file_content: bytes,
    filename: str,
    group_by_columns: list[str] | None = None,
    sort_by_column: str | None = None,
) -> dict[str, Any]:
    """
    Parse uploaded file (CSV or Excel) based on extension.

    When group_by_columns is provided, rows are aggregated into one row per
    unique combination of group_by_columns values. All other input_* columns
    are collected into an input_notes list within each row.
    """
    path = Path(filename)
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        parsed = parse_excel(file_content, filename)
    else:
        parsed = parse_csv(file_content, filename)

    if not group_by_columns:
        return parsed

    parsed["rows"] = group_rows(parsed["rows"], group_by_columns, sort_by_column)
    parsed["row_count"] = len(parsed["rows"])

    note_cols = [c for c in parsed["input_columns"] if c not in group_by_columns]
    grouped_input_cols = group_by_columns + (["input_notes"] if note_cols else [])
    parsed["input_columns"] = grouped_input_cols
    parsed["column_names"] = grouped_input_cols + parsed["output_columns"]

    return parsed
